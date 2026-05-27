from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
import pandas as pd
import math

wip_file='/home/kali/Downloads/Test2/CyberRani_WIP_May2026.xlsx'

targets = [
    ("Mini Gupta", "Partner", 34115.39),
    ("Harmanpreet Pall", "Senior Manager", 25216.51),
    ("Mohit Sisodiya", "Manager", 313055.06),
    ("Ankur Sharma", "Senior Consultant", 2675171.28),
    ("Varun Avinash Salunkhe", "Consultant", 2172978.81),
    ("Prerna Nahata", "Consultant", 1768833.43),
    ("Rahul Satyaprakash Rana", "Consultant", 1768833.43),
    ("Prithvi Singh", "Associate Consultant", 1541796.09),
]

# Load workbook
wb = load_workbook(wip_file, data_only=True)

# Designation mapping
pivot = wb['Pivot']
designation_map = {}

for r in range(1, pivot.max_row + 1):
    des = pivot.cell(r,1).value
    emp = pivot.cell(r,2).value
   
    if emp and des:
        designation_map[str(emp).strip()] = str(des).split('-')[-1].strip()

# Detail sheet
detail = wb['Detail']
rows = list(detail.values)

headers = rows[1]
df = pd.DataFrame(rows[2:], columns=headers)

emp_col = [c for c in df.columns if 'Employee' in str(c)][0]
date_col = [c for c in df.columns if 'Week Ending' in str(c)][0]
hour_col = [c for c in df.columns if 'Charged Hours' in str(c)][0]
ansr_col = [c for c in df.columns if 'ANSR' in str(c)][0]

df = df[[emp_col,date_col,hour_col,ansr_col]].copy()
df.columns = ['Employee','Week Ending Date','Hours','ANSR']

df = df[df['Employee'].notna()]
df = df[df['Hours'].notna()]
df = df[df['ANSR'].notna()]

df['Employee'] = df['Employee'].astype(str).str.strip()
df['Hours'] = pd.to_numeric(df['Hours'], errors='coerce')
df['ANSR'] = pd.to_numeric(df['ANSR'], errors='coerce')
df['Designation'] = df['Employee'].map(designation_map)

# NET by Employee + Week Ending Date (so negative adjustments cancel out)
df = (
    df.groupby(['Employee', 'Week Ending Date', 'Designation'], dropna=False)
    .agg({'Hours': 'sum', 'ANSR': 'sum'})
    .reset_index()
)

# Keep only rows with positive net hours and ANSR
df = df[(df['Hours'] > 0) & (df['ANSR'] > 0)]

df = df.reset_index(drop=True)
df['unique_id'] = df.index

# latest first
df = df.sort_values('Week Ending Date', ascending=False)

# totals
emp_totals = (
    df.groupby(['Designation','Employee'])['ANSR']
    .sum()
    .reset_index()
)

used_hours = {}  # key -> hours already consumed from that row

# Load previous output files to exclude already-reversed rows
import os
import glob
import re

previous_files = []
# Batch files only
batch_pattern = '/home/kali/Downloads/Test2/reversal_mapping_batch_*.xlsx'
for f in sorted(glob.glob(batch_pattern), key=lambda x: [int(c) if c.isdigit() else c for c in re.split(r'(\d+)', x)]):
    if os.path.exists(f):
        previous_files.append(f)

print(f"Scanning for previously used hours in: {[os.path.basename(f) for f in previous_files]}")

for prev_output in previous_files:
    try:
        wb_prev = load_workbook(prev_output, data_only=True)
        ws_prev = wb_prev.active
        prev_rows = list(ws_prev.values)
        if len(prev_rows) > 1:
            prev_df = pd.DataFrame(prev_rows[1:], columns=prev_rows[0])
            if 'Reversal From Employee' in prev_df.columns and 'Hours' in prev_df.columns:
                prev_df = prev_df[prev_df['Reversal From Employee'].notna()]
                prev_df = prev_df[prev_df['Hours'].notna()]
                prev_df['Hours'] = pd.to_numeric(prev_df['Hours'], errors='coerce')
                prev_df = prev_df[prev_df['Hours'] > 0]
                
                # Match previous allocations to current df unique_ids
                matched_count = 0
                for _, prev_row in prev_df.iterrows():
                    emp = str(prev_row['Reversal From Employee']).strip()
                    date_val = prev_row['Week Ending Date']
                    if hasattr(date_val, 'strftime'):
                        date_str = date_val.strftime('%d-%b-%Y')
                    else:
                        date_str = str(date_val).strip()
                    
                    hrs = float(prev_row['Hours'])
                    
                    # Find matching row in df
                    for _, df_row in df.iterrows():
                        d = df_row['Week Ending Date']
                        d_str = d.strftime('%d-%b-%Y') if hasattr(d, 'strftime') else str(d)
                        if df_row['Employee'] == emp and d_str == date_str:
                            key = df_row['unique_id']
                            used_hours[key] = used_hours.get(key, 0) + hrs
                            matched_count += 1
                            break
                print(f"Loaded {len(prev_df)} rows from {os.path.basename(prev_output)} (matched {matched_count} current source rows).")
        wb_prev.close()
    except Exception as e:
        print(f"Warning: Could not read {prev_output}: {e}")

print(f"Total unique source rows with pre-allocated hours: {len(used_hours)}")

output = []

# MAIN LOGIC
for target_emp, target_des, target_ansr in targets:
   
    remaining = target_ansr
   
    allowed = [target_des]
   
    if 'Senior Manager' in target_des:
        allowed.append('Manager')
   
    if 'Senior Analyst' in target_des:
        allowed.append('Associate Consultant')
        allowed.append('Consultant')

    if 'Associate Consultant' in target_des:
        allowed.append('Consultant')
   
    eligible = (
        emp_totals[emp_totals['Designation'].isin(allowed)]
        .sort_values('ANSR', ascending=False)
    )
   
    for _, emp_row in eligible.iterrows():
       
        rev_emp = emp_row['Employee']
       
        emp_rows = df[
            (df['Employee'] == rev_emp) &
            (df['Designation'].isin(allowed))
        ].copy()
       
        emp_rows = emp_rows.sort_values(
            'Week Ending Date',
            ascending=False
        )
       
        for _, row in emp_rows.iterrows():
           
            if remaining <= 0:
                break
           
            key = row['unique_id']
           
            total_hours = float(row['Hours'])
            total_ansr = round(float(row['ANSR']), 2)
            rate_per_hour = total_ansr / total_hours
           
            # How many hours are still available in this row?
            already_used = used_hours.get(key, 0)
            avail_hours = total_hours - already_used
           
            if avail_hours <= 0:
                continue
           
            avail_ansr = round(rate_per_hour * avail_hours, 2)
           
            # Partial allocation: take only what's needed
            if avail_ansr > remaining and avail_hours > 1:
                needed_hours = math.ceil(remaining / rate_per_hour)
                needed_hours = max(1, min(needed_hours, avail_hours))
                hours = needed_hours
                ansr = round(rate_per_hour * hours, 2)
            else:
                hours = avail_hours
                ansr = avail_ansr
           
            output.append({
                "Target Employee": target_emp,
                "Designation": target_des,
                "Target ANSR": round(target_ansr,2),
                "Reversal From Employee": rev_emp,
                "Reversal Designation": designation_map.get(rev_emp, ''),
                "Week Ending Date":
                    row['Week Ending Date'].strftime('%d-%b-%Y')
                    if hasattr(row['Week Ending Date'],'strftime')
                    else str(row['Week Ending Date']),
                "Hours": hours,
                "ANSR": ansr,
            })
           
            remaining -= ansr
           
            used_hours[key] = already_used + hours
       
        if remaining <= 0:
            break

out_df = pd.DataFrame(output)

# remove any accidental invalid rows
out_df = out_df[
    (out_df['Hours'] > 0) &
    (out_df['ANSR'] > 0)
]

# SAVE with precise formatting to match the reference workbook exactly
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
total_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

header_font = Font(name="Calibri", size=11, bold=True)
data_font = Font(name="Calibri", size=11, bold=False)
total_font = Font(name="Calibri", size=11, bold=True)

center_align = Alignment(horizontal="center", vertical="center")
left_top_align = Alignment(horizontal="left", vertical="top")
num_format_ansr = '_ * #,##0_ ;_ * \\-#,##0_ ;_ * "-"??_ ;_ @_ '

wb_out = Workbook()
ws = wb_out.active
ws.title = 'Clean Final Mapping'

# Headers matching the reference exactly
headers = ['Target Employee', 'Designation', 'Target ANSR', 'Reversal From Employee', 'Reversal Designation', 'Week Ending Date', 'Hours', 'ANSR']
ws.append(headers)

# Apply header styles
for col_idx in range(1, 9):
    cell = ws.cell(row=1, column=col_idx)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border
    if col_idx in [3, 8]:
        cell.number_format = num_format_ansr

grouped = out_df.groupby('Target Employee', sort=False)

for target_emp, group in grouped:
    start_row = ws.max_row + 1
    
    # Add data rows
    for idx, row in group.iterrows():
        row_data = [
            row['Target Employee'],
            row['Designation'],
            row['Target ANSR'],
            row['Reversal From Employee'],
            row['Reversal Designation'],
            row['Week Ending Date'],
            row['Hours'],
            row['ANSR']
        ]
        ws.append(row_data)
        curr_row = ws.max_row
        
        # Style the data row
        for col_idx in range(1, 9):
            c_cell = ws.cell(row=curr_row, column=col_idx)
            c_cell.font = data_font
            if col_idx in [1, 2, 3]:
                c_cell.alignment = left_top_align
            else:
                c_cell.alignment = center_align
            c_cell.border = thin_border
            if col_idx in [3, 8]:
                c_cell.number_format = num_format_ansr
                
    end_row = ws.max_row
    
    # Vertically merge columns A, B, C for this group if multiple rows exist
    if start_row < end_row:
        ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
        ws.merge_cells(start_row=start_row, start_column=2, end_row=end_row, end_column=2)
        ws.merge_cells(start_row=start_row, start_column=3, end_row=end_row, end_column=3)
        
    # Add Total row
    total_row_idx = end_row + 1
    total_ansr = round(group['ANSR'].sum(), 2)
    
    # Write values
    ws.cell(row=total_row_idx, column=1, value='Total')
    ws.cell(row=total_row_idx, column=8, value=total_ansr)
    
    # Style all cells in total row
    for col_idx in range(1, 9):
        t_cell = ws.cell(row=total_row_idx, column=col_idx)
        t_cell.font = total_font
        t_cell.alignment = center_align
        
        if col_idx == 1:
            t_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            t_cell.fill = total_fill
        elif col_idx == 8:
            t_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            t_cell.fill = total_fill
            t_cell.number_format = num_format_ansr
        else:
            t_cell.border = Border(top=Side(style='thin'), bottom=Side(style='thin'))
            if col_idx == 7:
                t_cell.border = Border(top=Side(style='thin'), bottom=Side(style='thin'), right=Side(style='thin'))
                
    # Merge A to G in the Total row
    ws.merge_cells(start_row=total_row_idx, start_column=1, end_row=total_row_idx, end_column=7)

# Set precise column widths to match reference exactly
column_widths = {
    'A': 20,
    'B': 20,
    'C': 16,
    'D': 22,
    'E': 20,
    'F': 17,
    'G': 10,
    'H': 15
}
for col_letter, width in column_widths.items():
    ws.column_dimensions[col_letter].width = width

# Auto-increment output filename
batch = 1
while os.path.exists(f'/home/kali/Downloads/Test2/reversal_mapping_batch_{batch}.xlsx'):
    batch += 1
out_path = f'/home/kali/Downloads/Test2/reversal_mapping_batch_{batch}.xlsx'
wb_out.save(out_path)

# Print summary
print(out_df.head(20))
print("\n--- ANSR Summary per Target Employee ---")
for target_emp, group in grouped:
    print(f"  {target_emp}: {round(group['ANSR'].sum(), 2)}")
print("\nSaved:", out_path)

